from win32com import client
from unittest import mock
# Set max font family value to 100
p = mock.patch('openpyxl.styles.fonts.Font.family.max', new=100)
p.start()

from tqdm import tqdm
from tqdm import TqdmWarning
import openpyxl
from openpyxl import load_workbook
import shutil
import time
import os
import time

class Autosave():
    def __init__(self,src,select,branch,mouth) -> None:
        self.branch = branch
        self.select = select
        self.mouth = mouth
        self.file = src
        self.savews = None
        self.temporaries = self.get_temporaries()
        self.sources = self.get_sources()["sources"]
        self.salib = self.get_sources()["salib"]
        self.temporary=self.get_sources()["temporary"]
        self.bar = None
        self.updatebar = None

    def calbar(self):
        percentage= 100
        target = 0
        if not self.select :
            sheets = []
            for sheet in self.sources:
                for i in range(self.get_round(sheet)):
                    target += 1
                    sheets.append(i)
            for i in sheets:
                target += 1
            result = (percentage/target)
            return result
        else:
            target = 2
            result = (percentage/target)
            return result

    def main(self):
        self.mkdir()
        self.updatebar = self.calbar()
        with tqdm(total=100, desc="Preparing", bar_format="{l_bar}{bar}|") as pbar:
            self.bar = pbar
            self.extract_convert(self.select)
        print("Successful")

    def get_temporaries(self):
        temporaries = 'temporary.xlsx'
        if os.path.exists(f"{temporaries}"):
            os.remove(f"{temporaries}")
        return temporaries

    def get_value(self,source,col,i):
        result = source.cell(row=i, column=col).value
        if result == None:
            result = "-"
        return result

    def get_round(self,source):
        i=0
        while source.cell(row=i+3, column=1).value != None:
            i+=1
        return i
    
    def get_sources(self):
        shutil.copyfile(self.file,self.temporaries)
        temporary = load_workbook(self.temporaries,data_only=True)
        salib = [ i for i in temporary if "สลิป" in i.title]
        sources = [ i for i in temporary if "สลิป" not in i.title]
        return {
            "sources":sources,
            "salib":salib,
            "temporary":temporary
            }

    def mkdir(self):
        for i in self.sources:
            path = f'{i.title}'
            if not os.path.exists(path):
                os.makedirs(path)


    def extract_convert(self,person=False):
        app = client.DispatchEx("Excel.Application")
        app.Interactive = False
        app.Visible = False
        if not person:
            for sheet in self.sources:
                for i in self.temporary.sheetnames:
                    if i != "สลิป "+sheet.title:
                        self.temporary.remove(self.temporary[i])
                file = []
                for i in range(self.get_round(sheet)):
                        img = openpyxl.drawing.image.Image('Please (15).png')
                        img.anchor = 'B1'
                        ws = [i.title for i in self.salib if i.title.split(" ")[1] == sheet.title]
                        salib = self.temporary[ws[0]]
                        salib.add_image(img)
                        i += 3
                        salib["C5"] = self.get_value(sheet,1,i) #รหัสพนักงาน
                        salib["C6"] = self.get_value(sheet,2,i) #ชื่อ-สกุล
                        salib["C7"] = self.get_value(sheet,4,i) #ตำเเหน่ง
                        salib["B8"] = f"ประจำเดือน {self.mouth}"
                        salib["C11"] = self.get_value(sheet,5,i) #อัตราเงินเดือน
                        salib["C12"] = self.get_value(sheet,6,i) #ค่าตำแหน่ง
                        salib["C13"] = self.get_value(sheet,7,i) #เบี้ยขยัน
                        salib["I11"] = self.get_value(sheet,8,i) #เบิก
                        salib["I12"] = self.get_value(sheet,9,i) #ประกันสังคม
                        salib["I14"] = self.get_value(sheet,10,i) #ยอดจ่ายเงินกู้
                        salib["I15"] = self.get_value(sheet,11,i) #ยอดเงินกู้คงเหลือ
                        salib["C17"] = self.get_value(sheet,12,i) #สวัสดิการอื่นๆ
                        salib["I13"] = self.get_value(sheet,13,i) #ขาด/สาย/ลา
                        salib["C14"] = self.get_value(sheet,14,i) #Incentive
                        salib["C16"] = self.get_value(sheet,15,i) #นักขัตฤกษ์
                        salib["C15"] = self.get_value(sheet,16,i) #ประเมิน
                        salib["C18"] = self.get_value(sheet,17,i) #โบนัส
                        salib["C22"] = self.get_value(sheet,18,i) #รายได้สุทธิ
                        salib["G16"] = self.get_value(sheet,20,i) #ขาด (วัน)
                        salib["G17"] = self.get_value(sheet,21,i) #สาย (นาที)
                        salib["G18"] = self.get_value(sheet,22,i) #ลาป่วย (วัน)
                        salib["I16"] = self.get_value(sheet,23,i) #ลากิจ (วัน)
                        salib["I17"] = self.get_value(sheet,24,i) #ลาพักร้อน (วัน)

                        salib["C20"] = '=SUM(C11:C19)' #รวมเงินได้
                        salib["I20"] = '=SUM(I11:I14)' #รวมรายการหัก
                        filename = f"{self.get_value(sheet,2,i+3)},{self.get_value(sheet,21,i+3)}"
                        self.temporary.save(f"{sheet.title}\\{filename}.xlsx")
                        file.append(salib['C6'].value)
                        self.bar.update(self.updatebar)
                        wb = app.Workbooks.Open(f"{os.getcwd()}\\{sheet.title}\\{filename}.xlsx")
                        wb.ActiveSheet.ExportAsFixedFormat(0,f"{os.getcwd()}\\{sheet.title}\\{filename}")
                        wb.Close()   
                        os.remove(f"{os.getcwd()}\{sheet.title}\{filename}.xlsx")
                        self.bar.desc = f"Extracting to PDF branch {sheet.title}"
                        self.bar.update(self.updatebar)
                        i+=1
                shutil.copyfile(self.file,self.temporaries)
                self.temporary = load_workbook(self.temporaries,data_only=True)
        else:
            shutil.copyfile(self.file,self.temporaries)
            self.temporary = load_workbook(self.temporaries,data_only=True)
            sheet = self.temporary[self.branch]
            for i in self.temporary.sheetnames:
                    if i != "สลิป "+self.branch:
                        self.temporary.remove(self.temporary[i])
            salib = self.temporary["สลิป "+self.branch]
            img = openpyxl.drawing.image.Image(f'{os.getcwd()}\\Please (15).png')
            img.anchor = 'B1'
            salib.add_image(img)
            i = int(self.select)+2
            salib["C5"] = self.get_value(sheet,1,i) #รหัสพนักงาน
            salib["C6"] = self.get_value(sheet,2,i) #ชื่อ-สกุล
            salib["C7"] = self.get_value(sheet,4,i) #ตำเเหน่ง
            salib["B8"] = f"ประจำเดือน {self.mouth}"
            salib["C11"] = self.get_value(sheet,5,i) #อัตราเงินเดือน
            salib["C12"] = self.get_value(sheet,6,i) #ค่าตำแหน่ง
            salib["C13"] = self.get_value(sheet,7,i) #เบี้ยขยัน
            salib["I11"] = self.get_value(sheet,8,i) #เบิก
            salib["I12"] = self.get_value(sheet,9,i) #ประกันสังคม
            salib["I14"] = self.get_value(sheet,10,i) #ยอดจ่ายเงินกู้
            salib["I15"] = self.get_value(sheet,11,i) #ยอดเงินกู้คงเหลือ
            salib["C17"] = self.get_value(sheet,12,i) #สวัสดิการอื่นๆ
            salib["I13"] = self.get_value(sheet,13,i) #ขาด/สาย/ลา
            salib["C14"] = self.get_value(sheet,14,i) #Incentive
            salib["C16"] = self.get_value(sheet,15,i) #นักขัตฤกษ์
            salib["C15"] = self.get_value(sheet,16,i) #ประเมิน
            salib["C18"] = self.get_value(sheet,17,i) #โบนัส
            salib["C22"] = self.get_value(sheet,18,i) #รายได้สุทธิ
            salib["G16"] = self.get_value(sheet,20,i) #ขาด (วัน)
            salib["G17"] = self.get_value(sheet,21,i) #สาย (นาที)
            salib["G18"] = self.get_value(sheet,22,i) #ลาป่วย (วัน)
            salib["I16"] = self.get_value(sheet,23,i) #ลากิจ (วัน)
            salib["I17"] = self.get_value(sheet,24,i) #ลาพักร้อน (วัน)

            salib["C20"] = '=SUM(C11:C19)' #รวมเงินได้
            salib["I20"] = '=SUM(I11:I14)' #รวมรายการหัก
            self.bar.desc = f"Extracting to Excel branch {sheet.title}"
            filename = f"{self.get_value(sheet,2,i+3)},{self.get_value(sheet,19,int(self.select)+2)}"
            self.temporary.save(f"{sheet.title}\\{filename}.xlsx")
            self.bar.update(self.updatebar)
            wb = app.Workbooks.Open(f"{os.getcwd()}\\{sheet.title}\\{filename}.xlsx")
            wb.ActiveSheet.ExportAsFixedFormat(0,f"{os.getcwd()}\\{sheet.title}\\{filename}")
            wb.Close()
            os.remove(f"{os.getcwd()}\\{sheet.title}\\{filename}.xlsx")
        self.bar.desc = f"Extraction complete {sheet.title}"
        self.bar.close
        os.remove(self.temporaries)



